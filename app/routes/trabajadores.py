from flask import Blueprint, render_template, request, redirect, url_for, flash
from flask_login import login_required, current_user
from app.models import Trabajador, Centro
from app import db
from flask_babel import gettext as _
from sqlalchemy.orm import joinedload
from sqlalchemy import or_
import re


bp = Blueprint('trabajadores', __name__, url_prefix='/trabajadores')

paises = ["España", "Francia", "Alemania", "Italia", "Portugal", "Pakistán", "SriLanka", "India", "Marruecos", "Bangladesh", "Ecuador", "Bolivia", "Argentina", "México", "Colombia", "Chile", "Perú"]

@bp.route('/', methods=['GET', 'POST'])
@login_required
def lista_trabajadores():
    if current_user.rol not in ['admin', 'empresa', 'responsable']:
        flash(_('No tienes permiso para ver esta página.'), 'danger')
        return redirect(url_for('dashboard.index'))

    ver_eliminados = request.args.get('eliminados') == '1'
    ver_huerfanos = request.args.get('huerfanos') == '1'
    centro_filtro_id = request.args.get('centro_filtro', type=int)

    
   # Redirigir a la vista independiente si se solicita ver huérfanos
    if ver_huerfanos and not ver_eliminados:
        from flask import get_flashed_messages
        get_flashed_messages(with_categories=True) 
        return redirect(url_for('trabajadores.trabajadores_sin_centro'))

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        dni = request.form.get('dni')
        nass = request.form.get('nass')
        nacionalidad = request.form.get('nacionalidad')
        genero = request.form.get('genero')
        iban = request.form.get('iban')
        email = request.form.get('email')
        telefono = request.form.get('telefono')
        centro_id = int(request.form.get('centro_id')) if request.form.get('centro_id') else None

        if not nombre or not email or not telefono or not dni or not centro_id:
            flash(_('Todos los campos son obligatorios.'), 'danger')
        else:
            # Comprobar duplicados
            duplicado = Trabajador.query.filter(
                or_(
                    Trabajador.dni == dni,
                    Trabajador.email == email,
                    Trabajador.telefono == telefono
                )
            ).first()

            if duplicado:
                if duplicado.dni == dni:
                    flash(_('Ya existe un trabajador con ese DNI/NIE.'), 'danger')
                elif duplicado.email == email:
                    flash(_('Ya existe un trabajador con ese email.'), 'danger')
                elif duplicado.telefono == telefono:
                    flash(_('Ya existe un trabajador con ese teléfono.'), 'danger')
            else:
                numeros = ''.join(re.findall(r'\d', dni))[:5]
                if len(numeros) < 5:
                    flash(_('El DNI/NIE debe contener al menos 5 números.'), 'danger')
                else:
                    nuevo = Trabajador(
                        nombre=nombre,
                        email=email,
                        telefono=telefono,
                        dni=dni,
                        centro_id=centro_id,
                        pin=numeros,
                        nass=nass,
                        nacionalidad=nacionalidad,
                        genero=genero,
                        iban=iban,
                        activo=True
                    )
                    db.session.add(nuevo)
                    db.session.commit()
                    flash(_('Trabajador creado correctamente.'), 'success')
                return redirect(url_for('trabajadores.lista_trabajadores'))


    query = Trabajador.query.options(joinedload(Trabajador.centro))

    if current_user.rol == 'admin':
        centros = Centro.query.all()
    elif current_user.rol == 'empresa':
        centros = Centro.query.filter_by(empresa_id=current_user.empresa_id).all()
    elif current_user.rol == 'responsable':
        centros = Centro.query.all()  # futuro: filtrar por centros asignados

    centros_ids = [c.id for c in centros]

    if current_user.rol != 'admin':
        query = query.filter(Trabajador.centro_id.in_(centros_ids))
    
    if centro_filtro_id:
        query = query.filter(Trabajador.centro_id == centro_filtro_id)


    # === NUEVO: lógica de huerfanos y eliminados ===
    if ver_huerfanos:
        centros_inactivos_ids = [c.id for c in Centro.query.filter_by(activo=False).all()]
        query = query.filter(
            Trabajador.activo == True,
            or_(
                Trabajador.centro_id == None,
                Trabajador.centro_id.in_(centros_inactivos_ids)
            )
        )
    elif ver_eliminados:
        query = query.filter(Trabajador.activo == False)
    else:
        query = query.filter(Trabajador.activo == True)


    sort = request.args.get('sort', 'id')
    order = request.args.get('order', 'asc')

    if sort == 'centro':
        query = query.join(Centro)
        sort_attr = Centro.nombre
    else:
        sort_attr = getattr(Trabajador, sort, Trabajador.id)

    sort_attr = sort_attr.desc() if order == 'desc' else sort_attr.asc()

    query = query.order_by(sort_attr)


    trabajadores = query.all()

    # === NUEVO: contador de huérfanos real ===
    centros_inactivos_ids = [c.id for c in Centro.query.filter_by(activo=False).all()]
    trabajadores_huerfanos = Trabajador.query.filter(
        Trabajador.activo == True,
        or_(
            Trabajador.centro_id == None,
            Trabajador.centro_id.in_(centros_inactivos_ids)
        )
    ).count()

    return render_template('trabajadores.html',
                           trabajadores=trabajadores,
                           centros=centros,
                           paises=paises,
                           ver_eliminados=ver_eliminados,
                           ver_huerfanos=ver_huerfanos,
                           trabajadores_huerfanos=trabajadores_huerfanos,
                           centro_filtro_id=centro_filtro_id)


@bp.route('/sin_centro')
@login_required
def trabajadores_sin_centro():
    if current_user.rol not in ['admin', 'empresa']:
        flash(_('No tienes permiso para ver esta página.'), 'danger')
        return redirect(url_for('dashboard.index'))

    # Limpiar mensajes flash anteriores
    from flask import get_flashed_messages
    _ = get_flashed_messages(with_categories=True)

    centros_inactivos_ids = [c.id for c in Centro.query.filter_by(activo=False).all()]
    query = Trabajador.query.options(joinedload(Trabajador.centro)).filter(
        Trabajador.activo == True,
        or_(
            Trabajador.centro_id == None,
            Trabajador.centro_id.in_(centros_inactivos_ids)
        )
    )

    # Ordenación por columna
    sort = request.args.get('sort', 'nombre')
    order = request.args.get('order', 'asc')

    if sort == 'centro':
        query = query.join(Centro)
        sort_attr = Centro.nombre
    else:
        sort_attr = getattr(Trabajador, sort, Trabajador.nombre)

    sort_attr = sort_attr.desc() if order == 'desc' else sort_attr.asc()
    query = query.order_by(sort_attr)

    trabajadores = query.all()


    centros = Centro.query.all()
    return render_template('trabajadores_sin_centro.html', trabajadores=trabajadores, centros=centros)


@bp.route('/reasignar/<int:trabajador_id>', methods=['POST'])
@login_required
def reasignar(trabajador_id):
    trabajador = Trabajador.query.get_or_404(trabajador_id)
    nuevo_centro_id = int(request.form.get('centro_id')) if request.form.get('centro_id') else None

    if not nuevo_centro_id:
        flash(_('Debes seleccionar un centro.'), 'danger')
        return redirect(url_for('trabajadores.trabajadores_sin_centro'))

    if current_user.rol == 'empresa':
        centro = Centro.query.get(nuevo_centro_id)
        if centro.empresa_id != current_user.empresa_id:
            flash(_('No puedes asignar a un centro que no pertenece a tu empresa.'), 'danger')
            return redirect(url_for('trabajadores.trabajadores_sin_centro'))

    trabajador.centro_id = nuevo_centro_id
    db.session.commit()
    flash(_('Trabajador reasignado correctamente.'), 'success')
    return redirect(url_for('trabajadores.trabajadores_sin_centro'))


@bp.route('/borrar/<int:trabajador_id>', methods=['POST'])
@login_required
def borrar(trabajador_id):
    trabajador = Trabajador.query.get_or_404(trabajador_id)
    trabajador.activo = False
    db.session.commit()
    flash(_('Trabajador eliminado lógicamente.'), 'warning')
    return redirect(url_for('trabajadores.lista_trabajadores', eliminados=0, huerfanos=0))


@bp.route('/restaurar/<int:trabajador_id>', methods=['POST'])
@login_required
def restaurar(trabajador_id):
    trabajador = Trabajador.query.get_or_404(trabajador_id)
    trabajador.activo = True
    db.session.commit()
    flash(_('Trabajador restaurado correctamente.'), 'success')
    return redirect(url_for('trabajadores.lista_trabajadores', eliminados=1))


@bp.route('/eliminar_definitivo/<int:trabajador_id>', methods=['POST'])
@login_required
def eliminar_definitivo(trabajador_id):
    trabajador = Trabajador.query.get_or_404(trabajador_id)
    db.session.delete(trabajador)
    db.session.commit()
    flash(_('Trabajador eliminado permanentemente.'), 'danger')
    return redirect(url_for('trabajadores.lista_trabajadores', eliminados=1))


@bp.route('/editar/<int:trabajador_id>', methods=['POST'])
@login_required
def editar(trabajador_id):
    trabajador = Trabajador.query.get_or_404(trabajador_id)

    nombre = request.form.get('nombre')
    dni = request.form.get('dni')
    nass = request.form.get('nass')
    nacionalidad = request.form.get('nacionalidad')
    genero = request.form.get('genero')
    iban = request.form.get('iban')
    email = request.form.get('email')
    telefono = request.form.get('telefono')
    centro_id = int(request.form.get('centro_id')) if request.form.get('centro_id') else None

    if not nombre or not email or not telefono or not dni or not centro_id:
        flash(_('Todos los campos son obligatorios.'), 'danger')
        return redirect(url_for('trabajadores.lista_trabajadores'))

    if trabajador.dni != dni:
        numeros = ''.join(re.findall(r'\d', dni))[:5]
        if len(numeros) < 5:
            flash(_('El nuevo DNI/NIE debe contener al menos 5 números.'), 'danger')
            return redirect(url_for('trabajadores.lista_trabajadores'))
        trabajador.pin = numeros

    trabajador.nombre = nombre
    trabajador.email = email
    trabajador.telefono = telefono
    trabajador.dni = dni
    trabajador.centro_id = centro_id
    trabajador.nass = nass
    trabajador.nacionalidad = nacionalidad
    trabajador.genero = genero
    trabajador.iban = iban

    db.session.commit()
    flash(_('Datos actualizados correctamente.'), 'success')
    return redirect(url_for('trabajadores.lista_trabajadores'))


@bp.route('/importar', methods=['GET', 'POST'])
@login_required
def importar_trabajadores():
    if current_user.rol not in ['admin', 'empresa']:
        flash(_('No tienes permiso para importar trabajadores.'), 'danger')
        return redirect(url_for('dashboard.index'))

    if request.method == 'POST':
        archivo = request.files.get('archivo')
        centro_id = request.form.get('centro_id')

        if not archivo or not centro_id:
            flash(_('Debes proporcionar un archivo y seleccionar un centro de trabajo.'), 'danger')
            return redirect(request.url)

        import openpyxl
        import re

        try:
            if archivo.filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(archivo)
                hoja = wb.active

                # Leer encabezados de la primera fila (los normalizamos a minúsculas)
                encabezados = [str(c.value).strip().lower() if c.value else "" for c in next(hoja.iter_rows(min_row=1, max_row=1))]

                requeridos = ['nombre', 'dni']
                insertados = 0
                duplicados = 0
                errores = 0

                for fila in hoja.iter_rows(min_row=2):
                    fila_dict = {
                        encabezados[i]: (cell.value if cell.value is not None else "").strip() if isinstance(cell.value, str)
                        else str(cell.value).strip() if cell.value is not None else ""
                        for i, cell in enumerate(fila)
                    }

                    if not all(fila_dict.get(campo, '').strip() for campo in requeridos):
                        errores += 1
                        continue

                    dni = fila_dict['dni'].strip()
                    if Trabajador.query.filter_by(dni=dni).first():
                        duplicados += 1
                        continue

                    pin = ''.join(re.findall(r'\d', dni))[:5]
                    if len(pin) < 5:
                        errores += 1
                        continue

                    nuevo = Trabajador(
                        nombre=fila_dict['nombre'].strip(),
                        dni=dni,
                        email=fila_dict.get('email') or None,
                        telefono=fila_dict.get('telefono') or None,
                        nass=fila_dict.get('nass') or None,
                        nacionalidad=fila_dict.get('nacionalidad') or None,
                        genero=fila_dict.get('genero') or None,
                        iban=fila_dict.get('iban') or None,
                        centro_id=int(centro_id),
                        pin=pin,
                        activo=True
                    )
                    db.session.add(nuevo)
                    insertados += 1

                db.session.commit()
                flash(_(f'{insertados} trabajadores importados. {duplicados} duplicados. {errores} con errores.'), 'success')
            else:
                flash(_('Formato de archivo no compatible. Usa archivo .xlsx'), 'danger')
        except Exception as e:
            flash(_('Error al procesar el archivo: ') + str(e), 'danger')

        return redirect(url_for('trabajadores.lista_trabajadores'))

    # GET
    if current_user.rol == 'admin':
        centros = Centro.query.all()
    else:
        centros = Centro.query.filter_by(empresa_id=current_user.empresa_id).all()

    return render_template('trabajadores_importar.html', centros=centros)




@bp.route('/descargar_plantilla_csv')
@login_required
def descargar_plantilla_csv():
    from flask import make_response
    import csv
    import io

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['nombre', 'dni', 'email', 'telefono', 'nass', 'nacionalidad', 'genero', 'iban'])
    output.seek(0)

    response = make_response(output.read())
    response.headers["Content-Disposition"] = "attachment; filename=plantilla_trabajadores.csv"
    response.headers["Content-type"] = "text/csv"
    return response
